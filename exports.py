"""ARAKONAK GES — Dışa aktarma motoru: Excel (.xlsx) ve PDF.

Streamlit'ten bağımsızdır; her iki fonksiyon da bytes döndürür.
Türkçe karakter desteği için DejaVuSans gömülüdür. PDF'e marka logosu
vektörel (svglib) olarak yerleştirilir; svglib yoksa metin başlığa düşer.
"""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import pandas as pd

import core

_HERE = Path(__file__).parent


def _asset(*names):
    """Dosyayı hem 'assets/…' klasöründe hem de kökte arar (düz yüklemeye dayanıklı)."""
    for n in names:
        for c in (_HERE / "assets" / "fonts" / n, _HERE / "assets" / n, _HERE / n):
            if c.exists():
                return c
    return _HERE / names[0]


FONT_REG = _asset("DejaVuSans.ttf")
FONT_BLD = _asset("DejaVuSans-Bold.ttf")
LOGO = _asset("logo.svg")

# Marka renkleri (turkuaz)
NAVY = "#0a6b76"; BLUE = "#0e7fb0"; GREEN = "#10b5a3"; RED = "#fb6f84"
AMBER = "#f2a93b"; INK = "#0b2a33"; GREY = "#7d9ea3"

# matplotlib Türkçe font
for _f in (FONT_REG, FONT_BLD):
    if _f.exists():
        fm.fontManager.addfont(str(_f))
plt.rcParams["font.family"] = "DejaVu Sans"


# ═══════════════════════════════ EXCEL ═══════════════════════════════
def build_excel(state: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df   = state["df"]; k = state["k"]; disc = state["disc"]
    gag  = state["gag"]; stock = state["stock"]; hse = state["hse"]
    meta = state["meta"]; delayed = state["delayed"]

    F = "Arial"
    thin = Side(style="thin", color="D6DBEC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="0B2A6B")
    hdr_font = Font(name=F, bold=True, color="FFFFFF", size=10)
    title_font = Font(name=F, bold=True, color="0B2A6B", size=16)
    label_font = Font(name=F, bold=True, color="28356B", size=10)
    val_font = Font(name=F, size=10, color="0B1437")

    wb = Workbook()

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def autofit(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── 1) ÖZET ──
    ws = wb.active; ws.title = "Özet"; ws.sheet_view.showGridLines = False
    ws["A1"] = f"{meta['name']} — Yönetici Özeti"; ws["A1"].font = title_font
    ws["A2"] = f"{meta['loc']}   ·   Rapor Tarihi: {date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = Font(name=F, size=10, color="6B76A8")

    kpi_rows = [
        ("Toplam Bütçe (BAC)", core.fmt_full(k["BAC"]), "$"),
        ("Kazanılan Değer (EV)", core.fmt_full(k["EV"]), "$"),
        ("Planlanan Değer (PV)", core.fmt_full(k["PV"]), "$"),
        ("Kalan İş", core.fmt_full(k["kalan"]), "$"),
        ("Fiziksel İlerleme", f"%{k['ilerleme']:.1f}", ""),
        ("Plana Göre Olması Gereken", f"%{k['planPct']:.1f}", ""),
        ("SPI (Zaman Performansı)", "—" if k["SPI"] is None else f"{k['SPI']:.3f}", ""),
        ("SV (Zaman Sapması)", core.fmt_full(k["SV"]), "$"),
        ("CPI (Maliyet Performansı)", "veri yok" if k["CPI"] is None else f"{k['CPI']:.3f}", ""),
        ("EAC (Tahmini Toplam Maliyet)", "veri yok" if k["EAC"] is None else core.fmt_full(k["EAC"]), "$"),
    ]
    r0 = 4
    ws.cell(r0, 1, "GÖSTERGE").font = label_font
    ws.cell(r0, 2, "DEĞER").font = label_font
    style_header(ws, r0, 2)
    for i, (lbl, val, _) in enumerate(kpi_rows, start=r0 + 1):
        ws.cell(i, 1, lbl).font = val_font
        c = ws.cell(i, 2, val); c.font = Font(name=F, bold=True, size=10, color="0B1437")
        c.alignment = Alignment(horizontal="right")
        for cc in (1, 2):
            ws.cell(i, cc).border = border

    sc = core.status_counts(df)
    sr = r0 + len(kpi_rows) + 2
    ws.cell(sr, 1, "DURUM DAĞILIMI").font = label_font
    for j, (s, n) in enumerate(sc, start=sr + 1):
        ws.cell(j, 1, s).font = val_font
        ws.cell(j, 2, n).font = Font(name=F, bold=True, size=10)
        for cc in (1, 2):
            ws.cell(j, cc).border = border
    autofit(ws, [34, 20])
    ws.cell(sr + len(sc) + 2, 1,
            "Not: Değerler uygulamadaki güncel duruma göre üretilmiştir. "
            "Maliyet göstergeleri yalnızca fiili maliyet (AC) girildiyse hesaplanır."
            ).font = Font(name=F, size=8, italic=True, color="8590BD")

    # ── 2) İŞ KALEMLERİ ──
    ws = wb.create_sheet("İş Kalemleri"); ws.sheet_view.showGridLines = False
    cols = [("ID", "id"), ("Grup", "grp"), ("Disiplin", "disc"), ("Poz Adı", "name"),
            ("Birim", "unit"), ("Miktar", "qty"), ("B.Fiyat ($)", "up"), ("Tutar ($)", "tutar"),
            ("Plan %", "plan"), ("Gerçek %", "real"), ("Kazanılan ($)", "comp"),
            ("Kalan ($)", "kalan"), ("Fiili Maliyet ($)", "ac"), ("Durum", "durum")]
    for j, (h, _) in enumerate(cols, start=1):
        ws.cell(1, j, h)
    style_header(ws, 1, len(cols))
    money_fmt = '#,##0'; pct_fmt = '0'
    dur_color = {"TAMAMLANDI": "E7F8EF", "DEVAM": "E8F0FF", "GERİDE": "FEECEC", "BAŞLAMADI": "F1F3FA"}
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        for j, (_, key) in enumerate(cols, start=1):
            val = row[key]
            cell = ws.cell(i, j, val)
            cell.font = Font(name=F, size=9); cell.border = border
            if key in ("qty", "up", "tutar", "comp", "kalan", "ac"):
                cell.number_format = money_fmt; cell.alignment = Alignment(horizontal="right")
            if key in ("plan", "real"):
                cell.number_format = pct_fmt; cell.alignment = Alignment(horizontal="center")
        dcell = ws.cell(i, len(cols))
        dcell.fill = PatternFill("solid", fgColor=dur_color.get(row["durum"], "FFFFFF"))
        dcell.alignment = Alignment(horizontal="center")
    # Toplam satırı — canlı SUM formülleri
    last = len(df) + 1; tr = last + 1
    ws.cell(tr, 4, "TOPLAM").font = Font(name=F, bold=True, size=10)
    for col_key, letter in (("tutar", "H"), ("comp", "K"), ("kalan", "L"), ("ac", "M")):
        c = ws.cell(tr, {"H": 8, "K": 11, "L": 12, "M": 13}[letter],
                    f"=SUM({letter}2:{letter}{last})")
        c.font = Font(name=F, bold=True, size=10); c.number_format = money_fmt
        c.fill = PatternFill("solid", fgColor="EEF2FB")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{last}"
    autofit(ws, [8, 11, 20, 52, 8, 12, 11, 13, 8, 9, 13, 13, 14, 12])

    # ── 3) DİSİPLİN ──
    ws = wb.create_sheet("Disiplin"); ws.sheet_view.showGridLines = False
    dcols = [("Disiplin", "disc"), ("Bütçe ($)", "budget"), ("Kazanılan ($)", "comp"),
             ("Plan %", "planPct"), ("Gerçek %", "realPct"), ("Sapma (puan)", "sapma")]
    for j, (h, _) in enumerate(dcols, start=1):
        ws.cell(1, j, h)
    style_header(ws, 1, len(dcols))
    for i, (_, row) in enumerate(disc.iterrows(), start=2):
        for j, (_, key) in enumerate(dcols, start=1):
            cell = ws.cell(i, j, round(float(row[key]), 1) if key != "disc" else row[key])
            cell.font = Font(name=F, size=9); cell.border = border
            if key in ("budget", "comp"):
                cell.number_format = money_fmt; cell.alignment = Alignment(horizontal="right")
            if "Pct" in key or key == "sapma":
                cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"; autofit(ws, [34, 15, 15, 10, 10, 13])

    # ── 4) GRUP ──
    ws = wb.create_sheet("Grup"); ws.sheet_view.showGridLines = False
    for j, h in enumerate(["Grup", "Bütçe ($)", "Kazanılan ($)", "Kalan ($)", "Plan %", "Gerçek %"], start=1):
        ws.cell(1, j, h)
    style_header(ws, 1, 6)
    for i, (_, row) in enumerate(gag.iterrows(), start=2):
        vals = [row["short"], row["budget"], row["comp"], row["kalan"],
                round(row["planPct"], 1), round(row["realPct"], 1)]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(i, j, v); cell.font = Font(name=F, size=9); cell.border = border
            if j in (2, 3, 4):
                cell.number_format = money_fmt
    autofit(ws, [12, 16, 16, 16, 10, 10])

    # ── 5) GECİKEN İŞLER ──
    ws = wb.create_sheet("Geciken İşler"); ws.sheet_view.showGridLines = False
    for j, h in enumerate(["Grup", "Disiplin", "Poz Adı", "Plan %", "Gerçek %",
                           "Gecikme (puan)", "Risk ($)"], start=1):
        ws.cell(1, j, h)
    style_header(ws, 1, 7)
    if delayed is not None and not delayed.empty:
        for i, (_, row) in enumerate(delayed.iterrows(), start=2):
            vals = [core.GROUP_SHORT.get(row["grp"], row["grp"]), row["disc"], row["name"],
                    round(row["plan"]), round(row["real"]), round(row["gecikme"]),
                    round(row["riskUSD"])]
            for j, v in enumerate(vals, start=1):
                cell = ws.cell(i, j, v); cell.font = Font(name=F, size=9); cell.border = border
                if j == 7:
                    cell.number_format = money_fmt
    else:
        ws.cell(2, 1, "Geride kalan kalem yok.").font = Font(name=F, italic=True, size=9)
    autofit(ws, [10, 20, 52, 8, 9, 13, 13])

    # ── 6) STOK ──
    ws = wb.create_sheet("Stok"); ws.sheet_view.showGridLines = False
    scols = list(stock.columns)
    headmap = {"name": "Malzeme", "unit": "Birim", "ordered": "Sipariş", "delivered": "Sevk",
               "onsite": "Sahada", "installed": "Montajlı", "remaining": "Kalan", "id": "ID"}
    for j, c in enumerate([headmap.get(x, x) for x in scols], start=1):
        ws.cell(1, j, c)
    style_header(ws, 1, len(scols))
    for i, (_, row) in enumerate(stock.iterrows(), start=2):
        for j, c in enumerate(scols, start=1):
            cell = ws.cell(i, j, row[c]); cell.font = Font(name=F, size=9); cell.border = border
    autofit(ws, [8, 34, 8] + [12] * (len(scols) - 3))

    # ── 7) İSG ──
    ws = wb.create_sheet("İSG"); ws.sheet_view.showGridLines = False
    for j, h in enumerate(["Gösterge", "Değer", "Birim"], start=1):
        ws.cell(1, j, h)
    style_header(ws, 1, 3)
    for i, (_, row) in enumerate(hse.iterrows(), start=2):
        for j, key in enumerate(["label", "value", "unit"], start=1):
            cell = ws.cell(i, j, row[key]); cell.font = Font(name=F, size=9); cell.border = border
    autofit(ws, [30, 16, 10])

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ═══════════════════════════════ PDF ═══════════════════════════════
def _png_group(gag: pd.DataFrame) -> bytes:
    fig, ax = plt.subplots(figsize=(5.4, 2.6), dpi=150)
    x = range(len(gag)); w = 0.38
    ax.bar([i - w/2 for i in x], gag["planPct"], w, label="Plan", color="#c9d8fb", edgecolor=BLUE)
    ax.bar([i + w/2 for i in x], gag["realPct"], w, label="Gerçek", color=GREEN)
    ax.set_xticks(list(x)); ax.set_xticklabels(gag["short"]); ax.set_ylim(0, 110)
    ax.set_ylabel("%"); ax.legend(fontsize=8, frameon=False, ncol=2)
    for i, (p, r) in enumerate(zip(gag["planPct"], gag["realPct"])):
        ax.text(i - w/2, p + 2, f"{p:.0f}", ha="center", fontsize=7, color=BLUE)
        ax.text(i + w/2, r + 2, f"{r:.0f}", ha="center", fontsize=7, color=GREEN)
    ax.spines[["top", "right"]].set_visible(False); ax.tick_params(labelsize=8)
    fig.tight_layout()
    b = io.BytesIO(); fig.savefig(b, format="png", transparent=True); plt.close(fig)
    return b.getvalue()


def _png_disc(disc: pd.DataFrame) -> bytes:
    d = disc.sort_values("budget", ascending=True).tail(8)
    fig, ax = plt.subplots(figsize=(5.4, 3.0), dpi=150)
    labels = [x if len(x) <= 22 else x[:20] + "…" for x in d["disc"]]
    ax.barh(labels, d["comp"], color=GREEN, label="Tamamlanan")
    ax.barh(labels, d["budget"] - d["comp"], left=d["comp"], color="#fbd5d1", label="Kalan")
    ax.set_xlabel("$"); ax.legend(fontsize=8, frameon=False, ncol=2)
    ax.spines[["top", "right"]].set_visible(False); ax.tick_params(labelsize=7)
    ax.xaxis.set_major_formatter(lambda v, _: f"{v/1e6:.1f}M" if v >= 1e6 else f"{v/1e3:.0f}K")
    fig.tight_layout()
    b = io.BytesIO(); fig.savefig(b, format="png", transparent=True); plt.close(fig)
    return b.getvalue()


def _png_scurve(baseline, snaps, ev_today) -> bytes:
    fig, ax = plt.subplots(figsize=(5.4, 2.6), dpi=150)
    if baseline is not None and not baseline.empty:
        ax.plot(baseline["date"], baseline["planPct"], "--", color=BLUE, lw=1.8, label="Plan baseline")
    if snaps is not None and not snaps.empty:
        ax.plot(snaps["date"], snaps["evPct"], "-o", color=GREEN, lw=2, ms=3, label="Gerçek (EV)")
    ax.plot([pd.Timestamp.today()], [ev_today], "o", color=GREEN, ms=9,
            markeredgecolor="white", markeredgewidth=1.5, label="Bugün")
    ax.set_ylim(0, 105); ax.set_ylabel("%"); ax.legend(fontsize=8, frameon=False)
    ax.spines[["top", "right"]].set_visible(False); ax.tick_params(labelsize=7)
    fig.autofmt_xdate(rotation=25)
    fig.tight_layout()
    b = io.BytesIO(); fig.savefig(b, format="png", transparent=True); plt.close(fig)
    return b.getvalue()


def build_pdf(state: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                                    Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    # Türkçe font — bulunamazsa Helvetica'ya düş (çökmez)
    try:
        if FONT_REG.exists() and FONT_BLD.exists():
            pdfmetrics.registerFont(TTFont("DejaVu", str(FONT_REG)))
            pdfmetrics.registerFont(TTFont("DejaVu-Bold", str(FONT_BLD)))
        else:
            raise FileNotFoundError
    except Exception:
        pdfmetrics.registerFont(pdfmetrics.Font("DejaVu", "Helvetica", "WinAnsiEncoding"))
        pdfmetrics.registerFont(pdfmetrics.Font("DejaVu-Bold", "Helvetica-Bold", "WinAnsiEncoding"))

    df = state["df"]; k = state["k"]; disc = state["disc"]; gag = state["gag"]
    meta = state["meta"]; delayed = state["delayed"]
    baseline = state.get("baseline"); snaps = state.get("snaps")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            title=f"{meta['name']} Rapor")
    W = doc.width
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontName="DejaVu-Bold",
                        textColor=colors.HexColor(NAVY), fontSize=15, spaceAfter=2)
    sub = ParagraphStyle("sub", fontName="DejaVu", fontSize=9, textColor=colors.HexColor("#6b76a8"))
    sec = ParagraphStyle("sec", fontName="DejaVu-Bold", fontSize=11,
                         textColor=colors.HexColor("#1e2d6a"), spaceBefore=10, spaceAfter=4)
    small = ParagraphStyle("small", fontName="DejaVu", fontSize=7.5,
                           textColor=colors.HexColor("#8590bd"))
    cellL = ParagraphStyle("cellL", fontName="DejaVu", fontSize=7.5, leading=9)

    elems = []

    # ── Başlık bandı (logo + isim) ──
    header_tbl_data = []
    try:
        from svglib.svglib import svg2rlg
        from reportlab.graphics import renderPDF  # noqa
        logo = svg2rlg(str(LOGO))
        target_h = 13 * mm
        scale = target_h / logo.height
        logo.width *= scale; logo.height *= scale; logo.scale(scale, scale)
        logo_flow = logo
    except Exception:
        logo_flow = Paragraph("<b>ARAKONAK</b>", ParagraphStyle(
            "lg", fontName="DejaVu-Bold", fontSize=16, textColor=colors.white))

    title_para = Paragraph(
        f"<font color='white'><b>{meta['name']}</b></font><br/>"
        f"<font color='#cfe0ff' size=8>Canlı İlerleme &amp; Bütçe — EPC Proje Kontrol Raporu</font>",
        ParagraphStyle("t", fontName="DejaVu-Bold", fontSize=14, leading=17))
    date_para = Paragraph(
        f"<font color='#cfe0ff' size=8>{meta['loc']}<br/>{date.today().strftime('%d.%m.%Y')}</font>",
        ParagraphStyle("d", fontName="DejaVu", fontSize=8, alignment=TA_RIGHT))

    band = Table([[logo_flow, title_para, date_para]],
                 colWidths=[34 * mm, W - 34 * mm - 30 * mm, 30 * mm])
    band.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(NAVY)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ROUNDEDCORNERS", [6, 6, 6, 6])]))
    elems += [band, Spacer(1, 8)]

    # ── KPI kutuları ──
    def kpi_cell(label, value, color=INK):
        return [Paragraph(f"<font color='#8590bd' size=7>{label}</font>", cellL),
                Paragraph(f"<font color='{color}' size=13><b>{value}</b></font>",
                          ParagraphStyle("kv", fontName="DejaVu-Bold", fontSize=13, leading=15))]

    spi = "—" if k["SPI"] is None else f"{k['SPI']:.2f}"
    spi_col = GREEN if (k["SPI"] and k["SPI"] >= 1) else RED
    kpi_data = [[
        kpi_cell("TOPLAM BÜTÇE", core.fmt_money(k["BAC"])),
        kpi_cell("KAZANILAN (EV)", core.fmt_money(k["EV"]), GREEN),
        kpi_cell("KALAN", core.fmt_money(k["kalan"]), RED),
        kpi_cell("İLERLEME", f"%{k['ilerleme']:.1f}", BLUE),
        kpi_cell("SPI", spi, spi_col),
    ]]
    flat = [[c[0] for c in kpi_data[0]], [c[1] for c in kpi_data[0]]]
    kpi_tbl = Table(flat, colWidths=[W / 5] * 5)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f6f8ff")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e7ecf9")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e7ecf9")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    elems += [kpi_tbl, Spacer(1, 4)]

    # durum satırı
    sc = core.status_counts(df)
    sc_txt = "   ".join([f"<font color='#1e2d6a'><b>{n}</b></font> {s}" for s, n in sc])
    elems += [Paragraph(sc_txt, ParagraphStyle("sc", fontName="DejaVu", fontSize=8,
                                                textColor=colors.HexColor("#7a86b8"))),
              Spacer(1, 6)]

    # ── Grafikler (2x1) ──
    elems += [Paragraph("Grup Bazlı İlerleme &nbsp;·&nbsp; S-Eğrisi", sec)]
    img_g = RLImage(io.BytesIO(_png_group(gag)), width=W / 2 - 4, height=(W / 2 - 4) * 0.48)
    img_s = RLImage(io.BytesIO(_png_scurve(baseline, snaps, k["ilerleme"])),
                    width=W / 2 - 4, height=(W / 2 - 4) * 0.48)
    row1 = Table([[img_g, img_s]], colWidths=[W / 2, W / 2])
    row1.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (0, 0), (-1, -1), 0),
                              ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    elems += [row1, Spacer(1, 4)]

    elems += [Paragraph("Disiplin Bazlı Tamamlanan / Kalan Bütçe", sec)]
    img_d = RLImage(io.BytesIO(_png_disc(disc)), width=W, height=W * 0.30)
    elems += [img_d, Spacer(1, 6)]

    # ── Geciken işler tablosu ──
    elems += [Paragraph("En Kritik Geciken İşler (bütçe ağırlıklı risk)", sec)]
    tdata = [["Grup", "Disiplin", "Poz", "Plan%", "Ger.%", "Risk $"]]
    if delayed is not None and not delayed.empty:
        for _, r in delayed.head(8).iterrows():
            tdata.append([
                core.GROUP_SHORT.get(r["grp"], r["grp"]), r["disc"][:16],
                Paragraph(r["name"][:70], cellL),
                f"{r['plan']:.0f}", f"{r['real']:.0f}", core.fmt_money(r["riskUSD"])])
    else:
        tdata.append(["—", "—", Paragraph("Geride kalan kalem yok.", cellL), "—", "—", "—"])
    dt = Table(tdata, colWidths=[16 * mm, 28 * mm, W - 16*mm - 28*mm - 16*mm - 16*mm - 20*mm,
                                 16 * mm, 16 * mm, 20 * mm])
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(NAVY)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "DejaVu"), ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVu-Bold"),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e0e5f2")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f8ff")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
    elems += [dt, Spacer(1, 8)]

    elems += [Paragraph(
        "Bu rapor uygulamadaki güncel duruma göre otomatik üretilmiştir. Maliyet göstergeleri "
        "(CPI/EAC) yalnızca fiili maliyet girildiyse hesaplanır; S-eğrisi baseline'ı "
        "proje başlangıç/bitiş tarihlerinden modellenmiştir.", small)]

    def footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("DejaVu", 7)
        canvas.setFillColor(colors.HexColor("#9aa4cc"))
        canvas.drawString(15 * mm, 8 * mm, f"{meta['name']} · Kontrol Panosu")
        canvas.drawRightString(doc_.pagesize[0] - 15 * mm, 8 * mm, f"Sayfa {doc_.page}")
        canvas.restoreState()

    doc.build(elems, onFirstPage=footer, onLaterPages=footer)
    return buf.getvalue()
